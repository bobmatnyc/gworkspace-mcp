"""Pandoc-based document conversion service.

Provides format detection and conversion between:
- Documents: docx, pdf, html, odt, rst <-> markdown
- Spreadsheets: xls, xlsx <-> csv, json  (via openpyxl, NOT pandoc)
- Presentations: pptx -> markdown (via pandoc)

Pandoc is invoked as an external subprocess (not via pypandoc) to keep the
dependency surface small and avoid version pinning issues.
"""

from __future__ import annotations

import csv
import io
import json
import subprocess  # nosec B404
import tempfile
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Format constants
# ---------------------------------------------------------------------------

# Maps file extension -> pandoc --from value
PANDOC_INPUT_FORMATS: dict[str, str] = {
    ".docx": "docx",
    ".doc": "docx",  # pandoc handles older .doc via docx
    ".odt": "odt",
    ".html": "html",
    ".htm": "html",
    ".rst": "rst",
    ".pdf": "pdf",  # limited support; pandoc can try
    ".pptx": "pptx",
    ".md": "markdown",
    ".txt": "plain",
    ".epub": "epub",
    ".latex": "latex",
    ".tex": "latex",
}

# Maps user-facing format name -> pandoc --to value
PANDOC_OUTPUT_FORMATS: dict[str, str] = {
    "md": "markdown",
    "markdown": "markdown",
    "html": "html",
    "rst": "rst",
    "docx": "docx",
    "odt": "odt",
    "pdf": "pdf",
    "latex": "latex",
    "txt": "plain",
}

# Spreadsheet formats handled by openpyxl (NOT pandoc)
SPREADSHEET_FORMATS: frozenset[str] = frozenset({".xls", ".xlsx", ".xlsm", ".ods"})

# Google Drive native MIME types -> rich export MIME types
# These allow downloading as Office formats that can then be converted.
GDRIVE_EXPORT_MIME: dict[str, str] = {
    # Google Docs -> DOCX -> pandoc -> markdown
    "application/vnd.google-apps.document": (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ),
    # Google Sheets -> XLSX -> openpyxl -> csv/json
    "application/vnd.google-apps.spreadsheet": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ),
    # Google Slides -> PPTX -> pandoc -> markdown
    "application/vnd.google-apps.presentation": (
        "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    ),
}

# Map export MIME type back to a file extension for temp-file naming
_MIME_TO_EXT: dict[str, str] = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": ".pptx",
}


# ---------------------------------------------------------------------------
# Exception
# ---------------------------------------------------------------------------


class ConversionError(Exception):
    """Raised when document conversion fails."""


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------


class PandocService:
    """First-class document conversion service using pandoc + openpyxl.

    Pandoc is invoked as a subprocess (never via pypandoc).  openpyxl is used
    for spreadsheet <-> CSV/JSON conversions so that Excel files are handled
    without requiring LibreOffice or pandoc.

    All public methods raise ``ConversionError`` on failure so callers can
    catch a single exception type regardless of the underlying tool.
    """

    # ------------------------------------------------------------------
    # Availability / version
    # ------------------------------------------------------------------

    def is_available(self) -> bool:
        """Return True if the pandoc binary is on PATH."""
        try:
            subprocess.run(  # nosec B603 B607
                ["pandoc", "--version"],
                capture_output=True,
                check=True,
                timeout=10,
            )
            return True
        except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired):
            return False

    def get_version(self) -> str | None:
        """Return the pandoc version string, or None if pandoc is not available."""
        try:
            result = subprocess.run(  # nosec B603 B607
                ["pandoc", "--version"],
                capture_output=True,
                check=True,
                text=True,
                timeout=10,
            )
            # First line is e.g. "pandoc 3.1.2"
            first_line = result.stdout.splitlines()[0] if result.stdout else ""
            return first_line.strip() or None
        except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired):
            return None

    # ------------------------------------------------------------------
    # Format detection
    # ------------------------------------------------------------------

    def detect_format(self, file_path: str | Path) -> str | None:
        """Return the pandoc input format name for *file_path*, or None.

        Detection is purely extension-based; no MIME sniffing is performed.
        Spreadsheet extensions that are handled by openpyxl return None here
        because they are not passed to pandoc.
        """
        ext = Path(file_path).suffix.lower()
        if ext in SPREADSHEET_FORMATS:
            return None  # Handled by openpyxl, not pandoc
        return PANDOC_INPUT_FORMATS.get(ext)

    # ------------------------------------------------------------------
    # Core conversion
    # ------------------------------------------------------------------

    def convert(
        self,
        input_path: str | Path,
        output_path: str | Path,
        from_format: str | None = None,
        to_format: str | None = None,
        extra_args: list[str] | None = None,
    ) -> Path:
        """Convert *input_path* to *output_path*.

        For spreadsheet inputs (.xls, .xlsx, .xlsm, .ods) the conversion is
        delegated to openpyxl rather than pandoc, producing CSV output.

        Args:
            input_path: Source file.
            output_path: Destination file.  Its extension is used to infer
                *to_format* when that argument is omitted.
            from_format: Pandoc ``--from`` value.  Auto-detected from
                *input_path* extension when None.
            to_format: Pandoc ``--to`` value or ``"csv"``/``"json"`` for
                spreadsheet targets.  Inferred from *output_path* extension
                when None.
            extra_args: Additional arguments passed verbatim to pandoc.

        Returns:
            Resolved ``Path`` to the output file.

        Raises:
            ConversionError: If pandoc is not installed, or if the conversion
                subprocess exits with a non-zero status.
        """
        input_path = Path(input_path)
        output_path = Path(output_path)

        in_ext = input_path.suffix.lower()
        out_ext = output_path.suffix.lower()

        # --- Spreadsheet branch (openpyxl) --------------------------------
        if in_ext in SPREADSHEET_FORMATS:
            resolved_to = (to_format or out_ext.lstrip(".") or "csv").lower()
            if resolved_to == "json":
                rows = self.spreadsheet_to_json(input_path)
                output_path.write_text(
                    json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
                )
            else:
                csv_text = self.spreadsheet_to_csv(input_path)
                output_path.write_text(csv_text, encoding="utf-8")
            return output_path.resolve()

        # --- CSV input -> spreadsheet branch (openpyxl) --------------------
        if in_ext == ".csv" and out_ext in (".xlsx", ".xls"):
            csv_text = input_path.read_text(encoding="utf-8")
            return self.csv_to_spreadsheet(csv_text, output_path)

        # --- Pandoc branch -------------------------------------------------
        self._require_pandoc()

        resolved_from = from_format or self.detect_format(input_path)
        resolved_to_raw = to_format or PANDOC_INPUT_FORMATS.get(out_ext, "markdown")
        # Normalise user-friendly names to pandoc format names
        resolved_to = PANDOC_OUTPUT_FORMATS.get(resolved_to_raw, resolved_to_raw)

        cmd: list[str] = ["pandoc", str(input_path), "-o", str(output_path)]
        if resolved_from:
            cmd += [f"--from={resolved_from}"]
        if resolved_to:
            cmd += [f"--to={resolved_to}"]
        if extra_args:
            cmd.extend(extra_args)

        try:
            subprocess.run(  # nosec B603 B607
                cmd,
                capture_output=True,
                check=True,
                text=True,
                timeout=120,
            )
        except subprocess.CalledProcessError as exc:
            raise ConversionError(
                f"pandoc conversion failed ({input_path.name} -> {output_path.name}): {exc.stderr}"
            ) from exc
        except subprocess.TimeoutExpired as exc:
            raise ConversionError(f"pandoc timed out converting {input_path.name}") from exc

        return output_path.resolve()

    def convert_bytes(
        self,
        content: bytes,
        from_format: str,
        to_format: str,
        filename_hint: str = "input",
    ) -> bytes:
        """Convert in-memory *content* bytes from one format to another.

        Writes *content* to a temporary file, runs ``convert()``, then reads
        the output back into memory.  Useful for stream-based pipelines such
        as download -> convert -> return.

        Args:
            content: Raw bytes of the source document.
            from_format: Pandoc ``--from`` value (e.g. ``"docx"``).
            to_format: Pandoc ``--to`` value or ``"csv"``/``"json"`` for
                spreadsheet targets.
            filename_hint: Base name used when constructing temporary file
                names (no path, no extension).

        Returns:
            Converted content as bytes.

        Raises:
            ConversionError: Propagated from :meth:`convert`.
        """
        # Determine extensions for the temp files
        in_ext = _pandoc_format_to_ext(from_format)
        out_ext = _pandoc_format_to_ext(to_format)

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            in_file = tmp / f"{filename_hint}{in_ext}"
            out_file = tmp / f"{filename_hint}_out{out_ext}"

            in_file.write_bytes(content)
            self.convert(in_file, out_file, from_format=from_format, to_format=to_format)
            return out_file.read_bytes()

    # ------------------------------------------------------------------
    # Spreadsheet helpers (openpyxl)
    # ------------------------------------------------------------------

    def spreadsheet_to_csv(self, input_path: Path) -> str:
        """Convert an xls/xlsx/ods file to a CSV string.

        Uses openpyxl to read the *first* worksheet.  Only the first sheet is
        converted; use :meth:`spreadsheet_to_json` for multi-sheet data.

        Args:
            input_path: Path to the spreadsheet file.

        Returns:
            CSV text (UTF-8, CRLF line endings per RFC 4180).

        Raises:
            ConversionError: If openpyxl is not installed or the file cannot
                be read.
        """
        wb = self._load_workbook(input_path)
        ws = wb.active
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):  # type: ignore[union-attr]
            writer.writerow([_cell_str(v) for v in row])
        return buf.getvalue()

    def spreadsheet_to_json(self, input_path: Path) -> list[dict[str, Any]]:
        """Convert an xls/xlsx/ods file to a list of row dicts.

        The first row is treated as the header.  All worksheets are included;
        each dict has a ``"_sheet"`` key with the sheet name.

        Args:
            input_path: Path to the spreadsheet file.

        Returns:
            List of row dicts, one per data row across all sheets.

        Raises:
            ConversionError: If openpyxl is not installed or the file cannot
                be read.
        """
        wb = self._load_workbook(input_path)
        rows: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
            if not all_rows:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(all_rows[0])]
            for data_row in all_rows[1:]:
                row_dict: dict[str, Any] = {"_sheet": sheet_name}
                for header, cell in zip(headers, data_row, strict=False):
                    row_dict[header] = _cell_value(cell)
                rows.append(row_dict)
        return rows

    def csv_to_spreadsheet(self, csv_content: str, output_path: Path) -> Path:
        """Convert a CSV string to an xlsx file using openpyxl.

        Args:
            csv_content: CSV text.
            output_path: Destination path (should end with ``.xlsx``).

        Returns:
            Resolved path to the created spreadsheet.

        Raises:
            ConversionError: If openpyxl is not installed or writing fails.
        """
        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError as exc:
            raise ConversionError(
                "openpyxl is required for spreadsheet conversion. "
                "Install it with: pip install openpyxl"
            ) from exc

        wb = openpyxl.Workbook()
        ws = wb.active
        reader = csv.reader(io.StringIO(csv_content))
        for row in reader:
            ws.append(row)  # type: ignore[union-attr]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            wb.save(str(output_path))
        except Exception as exc:
            raise ConversionError(f"Failed to save spreadsheet: {exc}") from exc
        return output_path.resolve()

    # ------------------------------------------------------------------
    # Helpers for the server's _upload_markdown_as_doc refactor
    # ------------------------------------------------------------------

    def markdown_to_docx(
        self,
        input_path: Path,
        output_path: Path,
        extra_args: list[str] | None = None,
    ) -> Path:
        """Convert a Markdown file to DOCX using pandoc.

        Thin wrapper around :meth:`convert` with explicit ``from``/``to``
        values so callers don't need to import format constants.

        Args:
            input_path: Path to source ``.md`` file.
            output_path: Destination ``.docx`` path.
            extra_args: Additional pandoc arguments.

        Returns:
            Resolved path to the output DOCX.
        """
        return self.convert(
            input_path,
            output_path,
            from_format="markdown",
            to_format="docx",
            extra_args=extra_args,
        )

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _require_pandoc(self) -> None:
        """Raise ConversionError with install instructions if pandoc is absent."""
        if not self.is_available():
            raise ConversionError(
                "pandoc is not installed. Install it with:\n"
                "  macOS:   brew install pandoc\n"
                "  Ubuntu:  sudo apt-get install pandoc\n"
                "  Windows: choco install pandoc\n"
                "  or see:  https://pandoc.org/installing.html"
            )

    def _load_workbook(self, path: Path) -> Any:
        """Load a workbook with openpyxl (read-only for performance)."""
        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError as exc:
            raise ConversionError(
                "openpyxl is required for spreadsheet conversion. "
                "Install it with: pip install openpyxl"
            ) from exc
        try:
            return openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise ConversionError(f"Failed to open spreadsheet {path.name}: {exc}") from exc


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------


def _pandoc_format_to_ext(fmt: str) -> str:
    """Return a file extension for a pandoc format name or user alias."""
    _map: dict[str, str] = {
        "docx": ".docx",
        "odt": ".odt",
        "html": ".html",
        "rst": ".rst",
        "markdown": ".md",
        "md": ".md",
        "pdf": ".pdf",
        "latex": ".tex",
        "plain": ".txt",
        "txt": ".txt",
        "pptx": ".pptx",
        "epub": ".epub",
        "csv": ".csv",
        "json": ".json",
        "xlsx": ".xlsx",
    }
    return _map.get(fmt.lower(), f".{fmt.lower()}")


def _cell_str(value: Any) -> str:
    """Convert a spreadsheet cell value to a string for CSV output."""
    if value is None:
        return ""
    return str(value)


def _cell_value(value: Any) -> Any:
    """Return a JSON-serialisable cell value for JSON output."""
    if value is None:
        return None
    # datetime / date objects: convert to ISO string
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value
